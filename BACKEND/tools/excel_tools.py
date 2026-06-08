"""
AERIS Excel & Screen Scraper Tools
Provides styled Excel sheet generation, smart context/role-based selection,
manual override support, and screen scraping/web search fallback capabilities.
"""
import os
import json
import logging
import asyncio
from pathlib import Path
from typing import Any, Dict, List, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import settings
from ai_engine import ai_engine

logger = logging.getLogger("aeris.tools.excel")


def get_backend_root() -> Path:
    """Helper to resolve the BACKEND directory root."""
    return Path(__file__).resolve().parent.parent


def get_workspace_dir() -> Path:
    """Get the active workspace directory path."""
    return Path(settings.WORKSPACE_DIR)


def format_excel_sheet(ws) -> None:
    """Apply premium styles to an openpyxl worksheet."""
    # Define styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3F66", end_color="1E3F66", fill_type="solid") # premium slate-blue
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="Calibri", size=11)
    cell_align_left = Alignment(horizontal="left", vertical="center")
    cell_align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border_side = Side(style="thin", color="CCCCCC")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    
    # 1. Format Headers (Row 1)
    ws.row_dimensions[1].height = 26
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
        
    # 2. Format Data Rows
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        # Zebra striping (light tint)
        fill_color = "F7F9FB" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.fill = row_fill
            cell.border = cell_border
            
            # Align based on data type
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = cell_align_right
            else:
                cell.alignment = cell_align_left

    # 3. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def export_to_excel(file_path: str, data: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> str:
    """
    Export a list of dicts to a styled Excel sheet.
    """
    try:
        workspace = get_workspace_dir()
        # Ensure path is workspace-relative or absolute within boundary
        resolved_path = Path(file_path)
        if not resolved_path.is_absolute():
            resolved_path = workspace / file_path
        
        # Ensure parent dirs exist
        resolved_path.parent.mkdir(parents=True, exist_ok=True)

        df = pd.DataFrame(data)
        if columns:
            # Reorder/filter columns to match requested
            existing_cols = [c for c in columns if c in df.columns]
            df = df[existing_cols]

        df.to_excel(resolved_path, index=False, engine="openpyxl")

        # Load workbook to apply styles
        wb = load_workbook(resolved_path)
        for ws in wb.worksheets:
            format_excel_sheet(ws)
        wb.save(resolved_path)
        wb.close()

        logger.info(f"Excel sheet successfully created: {resolved_path}")
        from utils.file_tracker import record_file_creation
        record_file_creation(str(resolved_path), f"Exported styled Excel sheet with {len(data)} rows")
        return f"Excel sheet successfully created at: {resolved_path}"
    except Exception as e:
        logger.error(f"Error in export_to_excel: {e}")
        raise RuntimeError(f"Failed to export to Excel: {e}")


async def update_excel_from_screen(
    excel_path_or_keyword: Optional[str] = None,
    target_name: Optional[str] = None,
    manual_details: Optional[Any] = None,
    source: Optional[str] = None
) -> str:
    """
    Updates or appends a person's details to an Excel sheet.
    - If manual_details is provided, writes them directly.
    - If not, captures screen and extracts details for target_name if visible.
    - If not visible/not found, falls back to web research/search ONLY if source is "web".
    - Otherwise, stops and asks the user for confirmation/details.
    - Smart naming: If no file is specified, maps the person's role to a sheet name.
    """
    workspace = get_workspace_dir()
    details: Dict[str, Any] = {}
    name_to_use = target_name or "Unknown"

    # Step 1: Resolve manual details if provided
    if manual_details:
        if isinstance(manual_details, dict):
            details = dict(manual_details)
        elif isinstance(manual_details, str):
            # Parse text details using LLM
            prompt = (
                f"Extract structured profile details from this text and return a JSON object.\n"
                f"Fields should include: Name, Role, Email, Phone, Company, Details.\n"
                f"Text: \"{manual_details}\""
            )
            try:
                extracted = await ai_engine.chat([
                    {"role": "system", "content": "You are a precise JSON extractor. Respond ONLY with valid JSON mapping fields."},
                    {"role": "user", "content": prompt}
                ], response_format={"type": "json_object"})
                details = json.loads(extracted.strip())
            except Exception as e:
                logger.warning(f"Failed to parse manual details text: {e}")
                details = {"Name": name_to_use, "Details": manual_details}
        
        if "Name" in details and details["Name"]:
            name_to_use = details["Name"]
        else:
            details["Name"] = name_to_use

    # Step 2: Screen Scraping if no manual details
    if not details and target_name and source != "web":
        try:
            import pyautogui
            screenshot = pyautogui.screenshot()
            screenshots_dir = get_backend_root() / "data"
            screenshots_dir.mkdir(parents=True, exist_ok=True)
            img_path = screenshots_dir / "vision_screen.png"
            screenshot.save(img_path)

            # Convert to base64
            import base64
            with open(img_path, "rb") as f:
                img_b64 = base64.b64encode(f.read()).decode("utf-8")

            prompt = (
                f"Examine this screen screenshot with extreme detail.\n"
                f"Identify if a person named '{target_name}' (or matching names, case-insensitive, e.g. '{target_name}') is visible anywhere on the screen (look inside chats, profiles, tables, document text, email threads, etc.).\n"
                f"If visible, extract EVERY single detail associated with them (such as Designation/Role, Email, Phone, Company, Status, Age, Location, etc.).\n"
                f"Return ONLY a valid JSON object with the exact fields:\n"
                f"{{\n"
                f"  \"found\": true,\n"
                f"  \"Name\": \"{target_name}\",\n"
                f"  \"Role\": \"extracted role\",\n"
                f"  \"Email\": \"extracted email\",\n"
                f"  \"Phone\": \"extracted phone\",\n"
                f"  \"Company\": \"extracted company\",\n"
                f"  \"Details\": \"other extracted details\"\n"
                f"}}\n"
                f"If not visible or not found, return {{\"found\": false}}"
            )
            vision_res = await ai_engine.vision(prompt, img_b64)
            vision_data = json.loads(vision_res.strip().strip("```json").strip("```").strip())
            if vision_data.get("found"):
                details = {k: v for k, v in vision_data.items() if k != "found"}
                logger.info(f"Successfully scraped details for {target_name} from screen.")
        except Exception as e:
            logger.warning(f"Screen scraping failed/unsupported: {e}")

    # Step 3: Web Search Fallback only if explicitly requested or forced
    if not details and target_name:
        if source != "web":
            logger.info(f"Details for '{target_name}' not found on screen. Asking user for confirmation.")
            return f"Sir, mujhe screen par '{target_name}' ki details nahi mili. Kya main in details ke liye web search karoon, ya aap details chat me provide karenge?"

        logger.info(f"Details not found on screen. Falling back to internet search for '{target_name}' as requested.")
        search_query = f"{target_name} contact email role designation profile organization"
        try:
            # Call realtime search directly
            from services.chat_engine import realtime_search
            # run in executor since realtime_search might be sync
            loop = asyncio.get_event_loop()
            search_result = await loop.run_in_executor(None, realtime_search, search_query)
            
            prompt = (
                f"Based on the following search result for '{target_name}', synthesize profile details.\n"
                f"Respond with ONLY JSON:\n"
                f"{{\n"
                f"  \"Name\": \"{target_name}\",\n"
                f"  \"Role\": \"Role/Designation\",\n"
                f"  \"Email\": \"Email or 'Not found'\",\n"
                f"  \"Phone\": \"Phone or 'Not found'\",\n"
                f"  \"Company\": \"Company or Organization\",\n"
                f"  \"Details\": \"Summary of profile\"\n"
                f"}}\n\n"
                f"Search results:\n{search_result}"
            )
            extracted = await ai_engine.chat([
                {"role": "system", "content": "You are a precise JSON extractor. Respond ONLY with valid JSON."},
                {"role": "user", "content": prompt}
            ], response_format={"type": "json_object"})
            details = json.loads(extracted.strip())
        except Exception as e:
            logger.warning(f"Web search fallback failed: {e}")
            details = {
                "Name": target_name,
                "Role": "Unknown",
                "Email": "Not found",
                "Phone": "Not found",
                "Company": "Unknown",
                "Details": "Failed to retrieve online details"
            }

    # Ensure a basic Name exists
    if not details:
        details = {
            "Name": name_to_use,
            "Role": "General",
            "Email": "Not specified",
            "Details": "No details could be gathered"
        }

    # Ensure Name column contains target_name
    if "Name" not in details or not details["Name"]:
        details["Name"] = name_to_use

    # Step 4: Role-based sheet resolving
    resolved_file_name = ""
    role = str(details.get("Role", "")).lower()
    designation = str(details.get("Designation", "")).lower()

    if excel_path_or_keyword:
        kw = excel_path_or_keyword.strip().lower()
        if kw.endswith(".xlsx"):
            resolved_file_name = excel_path_or_keyword
        else:
            resolved_file_name = f"{kw}.xlsx"
    else:
        # Smart routing based on role/designation
        if any(w in role or w in designation for w in ("hr", "human resource", "recruiter")):
            resolved_file_name = "hr.xlsx"
        elif any(w in role or w in designation for w in ("developer", "engineer", "coder", "programmer", "architect", "tech")):
            resolved_file_name = "developer.xlsx"
        elif any(w in role or w in designation for w in ("manager", "lead", "director", "vp", "ceo", "cfo", "cto")):
            resolved_file_name = "manager.xlsx"
        else:
            resolved_file_name = "details.xlsx"

    # Resolve relative to workspace
    file_path = Path(resolved_file_name)
    if not file_path.is_absolute():
        file_path = workspace / resolved_file_name

    file_path.parent.mkdir(parents=True, exist_ok=True)

    # Step 5: Incremental Update (Load, check name, update or append, save)
    try:
        if file_path.exists() and file_path.stat().st_size > 0:
            df = pd.read_excel(file_path, engine="openpyxl")
            # standardise columns
            df.columns = [str(c).strip() for c in df.columns]
            
            # Find name column (look for 'Name' or 'name')
            name_col = next((c for c in df.columns if c.lower() == "name"), None)
            if not name_col:
                name_col = "Name"
                df["Name"] = ""

            # Ensure all keys in details are columns in DataFrame
            for key in details.keys():
                if key not in df.columns:
                    df[key] = ""

            # Check if person exists (case-insensitive)
            name_match_idx = df[df[name_col].astype(str).str.lower() == details["Name"].lower()].index
            
            if not name_match_idx.empty:
                # Update existing row
                idx = name_match_idx[0]
                for key, val in details.items():
                    df.at[idx, key] = val
                logger.info(f"Updated existing entry for '{details['Name']}' in {file_path}")
            else:
                # Append new row
                new_row = {col: "" for col in df.columns}
                for key, val in details.items():
                    new_row[key] = val
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                logger.info(f"Appended new entry for '{details['Name']}' in {file_path}")
        else:
            # Create new sheet
            df = pd.DataFrame([details])
            logger.info(f"Created new Excel file at {file_path}")

        # Write and format
        df.to_excel(file_path, index=False, engine="openpyxl")
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            format_excel_sheet(ws)
        wb.save(file_path)
        wb.close()

        from utils.file_tracker import record_file_creation
        record_file_creation(str(file_path), f"Updated '{details.get('Name')}' details in Excel sheet")

        return f"Sir, I have successfully updated '{details['Name']}' details in the Excel sheet at: `{file_path}`"
    except Exception as e:
        logger.error(f"Error updating Excel sheet: {e}")
        raise RuntimeError(f"Failed to update Excel sheet: {e}")
